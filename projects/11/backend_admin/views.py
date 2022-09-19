import datetime
import time

from django.conf import settings
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.models import User
from django.core.handlers.wsgi import WSGIRequest
from django.shortcuts import render, redirect, HttpResponse
import openpyxl
import os

from django.urls import reverse
from django.views import View
from rest_framework import status
from rest_framework.response import Response

from backend_admin import forms
from backend_admin import models
from backend_api import models as models_api
from backend_api import serializers


class Logger:
    def __init__(self, request, error='-'):
        self.request = request
        self.error = error

    def log(self):
        try:
            if settings.DEBUG:
                print(f"error: {self.error}")
            if settings.ERROR_LOGGING:
                models.Log.objects.create(
                    path=self.request.path,
                    method=self.request.method,
                    user=self.request.user,
                    error=f"{self.error[:250]}",
                )
            return render(self.request, "components/404.html", context={})
        except Exception as error:
            Logger.log_static_error_txt(error=f"{error}")
            return render(self.request, "components/404.html", context={})

    @staticmethod
    def log_static(path: str, method: str) -> None:
        try:
            models.Log.objects.create(
                path=path,
                method=method
            )
        except Exception as error:
            Logger.log_static_error_txt(error=f"{error}")

    @staticmethod
    def log_static_error(error: str) -> None:
        try:
            models.Log.objects.create(
                error=error,
            )
        except Exception as error:
            Logger.log_static_error_txt(error=f"{error}")

    @staticmethod
    def log_static_error_txt(error: str) -> None:
        with open('static/admin/logs/log_error.txt', 'w') as file:
            file.write(f"{error} \t {datetime.datetime.now()}\n")


def staff_access_required(func):
    def wrap(*args, **kwargs):  # (admin, 12345qwerty,) {"username": 'admin', "password": '12345qwerty'}

        # if isinstance(kwargs.get("request", None), WSGIRequest):
        #     request = kwargs.get("request", None)
        # else:
        #     request = None

        if isinstance(args[0], WSGIRequest):
            request = args[0]
        elif isinstance(args[1], WSGIRequest):
            request = args[1]
        else:
            request = None

        if not request.user.is_authenticated:
            return redirect(reverse('login', args=()))
        elif not request.user.is_superuser and not request.user.is_staff:
            return redirect(reverse('login', args=()))

        result = func(*args, **kwargs)
        return result

    return wrap


def loger_action(func):
    def wrap(*args, **kwargs):
        if settings.ACTION_LOGGING:

            if isinstance(args[0], WSGIRequest):
                request = args[0]
            elif isinstance(args[1], WSGIRequest):
                request = args[1]
            else:
                request = None

            models.Log.objects.create(
                path=request.path,
                method=request.method,
                user=request.user,
                error="-",
            )
        result = func(*args, **kwargs)
        return result

    return wrap


class HomeView(View):
    try:
        pass
    except Exception as error:
        Logger.log_static_error(
            error=f"{error[:250]}"
        )

    @loger_action
    @staff_access_required
    def get(self, request):
        return render(request, 'backend_admin/Home.html', context={})


# def home(request):
#     return render(request, 'backend_admin/Home.html', context={})


class GetActiveUserListView(View):
    initial = {'key': 'value'}
    template_name = 'backend_admin/GetUserList.html'

    @staff_access_required
    def get(self, request, *args, **kwargs):
        users = User.objects.filter(is_active=True)  # только активные пользователи

        # EXCEL GENERATOR
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        kwargs_list = ["username", "password", "id", "is_superuser"]
        users_list = [x for x in users]  # list comprehension (pre generator yeld)
        for kwarg in kwargs_list:
            worksheet.cell(row=1, column=kwargs_list.index(kwarg) + 1, value=kwarg)
        for user_obj in users_list:
            row_index = users_list.index(user_obj)
            for kwarg in kwargs_list:
                col_index = kwargs_list.index(kwarg)
                worksheet.cell(row=row_index + 2, column=col_index + 1, value=getattr(user_obj, kwarg))
        path = f'temp/new_{datetime.datetime.now().strftime("%m-%d-%Y %H-%M-%S-%f")}.xlsx'
        directory = os.path.join("static", "temp")
        if not os.path.exists(directory):
            os.mkdir(directory)
        if os.path.exists(settings.STATIC_URL[1:] + path):
            os.remove(settings.STATIC_URL[1:] + path)
        workbook.save(settings.STATIC_URL[1:] + path)
        # EXCEL GENERATOR

        context = {
            "title": "Абракадабра",
            "headers": ["id", "username", "password", "password", "password", "is_superuser"],
            "user_list": users,
            "excel_file": path
        }
        return render(request, self.template_name, context)


from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import AllowAny


# @staff_access_required
@api_view(http_method_names=["GET", "POST"])
@permission_classes([AllowAny])
def get_active_user_list(request):  # функция-контроллер
    try:

        users = User.objects.filter(is_active=True)  # только активные пользователи

        # TODO EXCEL GENERATOR
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        kwargs_list = ["username", "password", "id", "is_superuser"]
        users_list = [x for x in users]  # list comprehension (pre generator yeld)
        for kwarg in kwargs_list:
            worksheet.cell(row=1, column=kwargs_list.index(kwarg) + 1, value=kwarg)
        for user_obj in users_list:
            row_index = users_list.index(user_obj)
            for kwarg in kwargs_list:
                col_index = kwargs_list.index(kwarg)
                worksheet.cell(row=row_index + 2, column=col_index + 1, value=getattr(user_obj, kwarg))
        path = f'temp/new_{datetime.datetime.now().strftime("%m-%d-%Y %H-%M-%S-%f")}.xlsx'
        directory = os.path.join("static", "temp")
        if not os.path.exists(directory):
            os.mkdir(directory)
        if os.path.exists(settings.STATIC_URL[1:] + path):
            os.remove(settings.STATIC_URL[1:] + path)
        workbook.save(settings.STATIC_URL[1:] + path)
        # TODO EXCEL GENERATOR

        context = {
            "title": "Абракадабра",
            "headers": ["id", "username", "password", "password", "password", "is_superuser"],
            "user_list": users,
            "excel_file": path
        }

        ser_users = serializers.UserSerializer(instance=users, many=True).data
        return Response(data={"user_list": ser_users}, status=status.HTTP_200_OK)

        # return render(request, 'backend_admin/GetUserList.html', context=context)
    except Exception as error:
        return Logger(request=request, error=str(error)).log()


def django_login(request):
    try:
        if request.method == "GET":
            context = {
                "login_form": forms.LoginUserForm()
            }
            return render(request, "backend_admin/Login.html", context=context)
        if request.method == "POST":

            print(request.POST)

            email = request.POST["email"]
            password = request.POST["password"]
            if email and password:
                user = authenticate(username=email, password=password)
                if user:
                    login(request, user)  # сессия сохраняется
                    return redirect(reverse('home', args=()))
            return redirect(reverse('login', args=()))
    except Exception as error:
        return Logger(request=request, error=str(error)).log()


def django_logout(request):
    try:
        logout(request)
        return redirect(reverse('login', args=()))
    except Exception as error:
        return Logger(request=request, error=str(error)).log()


@loger_action
@staff_access_required
def receipt(request):
    try:
        if request.method == "GET":
            categories = models_api.ReceiptCategory.objects.all()
            ingredients = models_api.ReceiptIngredient.objects.all()
            context = {"categories": categories, "ingredients": ingredients}
            return render(request, "backend_admin/CreateReceipt.html", context=context)
        elif request.method == "POST":
            pass
        elif request.method == "PUT" or request.method == "PATCH":
            pass
        elif request.method == "DELETE":
            pass
        else:
            return HttpResponse(status=405)
    except Exception as error:
        return Logger(request=request, error=str(error)).log()
