import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Border, PatternFill, GradientFill, Alignment, Side

# workbook = Workbook()  # для создания в памяти экземпляра Workbook
# worksheet = workbook.active

workbook = openpyxl.load_workbook('users.xlsx')
worksheet = workbook.active

#            0    1
new_list = [123, 124]
print(new_list[0])

new_dict = {"first": 123, 23: 666, "23": 666}
print(new_dict["first"])

value = worksheet["A56"].value
print(value)
print(type(value))

new_set = {12, 135, 656, 12}
print(new_set)
print(type(new_set))

new_tuple = (12,)

workers = []
for row in range(1, 2150):
    coordinates = "A" + str(row)  # 1 способ сложения строк
    # print(coordinates)
    coordinates = f"A{row}"  # 2 способ сложения строк
    # print(coordinates)
    coordinates = "A{}".format(row)  # 3 способ сложения строк
    # print(coordinates)
    col = "A"
    coordinates = "{col}{row}".format(col=col, row=row)  # 4 способ сложения строк
    # print(coordinates)
    # txt1 = "My name is {fname}, I'm {age}".format(fname="John", age=36)
    # txt2 = "My name is {0}, I'm {1}".format("John", 36)
    # txt3 = "My name is {}, I'm {}".format("John", 36)

    worker = []
    for col in "ABCDEF":
        coordinates = "{col}{row}".format(col=col, row=row)  # 4 способ сложения строк
        value = worksheet[coordinates].value
        if value is not None:
            # print(value)
            worker.append(value)
            pass
    workers.append(worker)


# print(workers)


class Worker:
    def __init__(self, first_name: str, second_name, patronymic, id, position, category):
        self.first_name = first_name
        self.second_name = second_name
        self.patronymic = patronymic
        self.id = id
        self.position = position
        self.category = category

    def get_full_name(self):
        return f"{self.first_name} {self.second_name}"


max_row = worksheet.max_row + 1
max_column = worksheet.max_column + 1

workers = []
for row in range(2, max_row):
    worker = []
    for column in range(1, max_column):
        value = worksheet.cell(row=row, column=column).value
        worker.append(value)
    # print(worker)
    # print(type(worker))
    worker = Worker(
        first_name=worker[0],
        second_name=worker[1],
        patronymic=worker[2],
        id=worker[3],
        position=worker[4],
        category=worker[5]
    )
    workers.append(worker)
    # print(worker)
    # print(type(worker))

print(workers[2])

workbook = Workbook()  # для создания в памяти экземпляра Workbook
worksheet = workbook.active

worksheet[f"A1"] = "Фамилия Имя"
worksheet.merge_cells('B2:C6')
a1 = worksheet[f"A1"]
a1.font = Font(color="00FF6600", bold=True, sz=28)

index = 1
for worker in workers:
    index += 1
    worksheet[f"A{index}"] = worker.get_full_name()

workbook.save('new1.xlsx')

wb = Workbook()
ws = wb.active
ws.merge_cells('B2:F4')
top_left_cell = ws['B2']
top_left_cell.value = "My Cell"
thin = Side(border_style="thin", color="000000")
double = Side(border_style="double", color="ff0000")
top_left_cell.border = Border(top=double, left=thin, right=thin, bottom=double)
top_left_cell.fill = PatternFill("solid", fgColor="DDDDDD")
top_left_cell.fill = fill = GradientFill(stop=("000000", "FFFFFF"))
top_left_cell.font  = Font(b=True, color="FF0000")
top_left_cell.alignment = Alignment(horizontal="center", vertical="center")
wb.save("styled.xlsx")